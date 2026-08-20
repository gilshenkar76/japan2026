"""Generate 'הצעות לשיפור קיוטו' as Excel and PDF files."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
import os
import arabic_reshaper
from bidi.algorithm import get_display

# ── Hebrew RTL helper ──────────────────────────────────────────────────────────
def h(text):
    """Prepare Hebrew/RTL text for ReportLab."""
    try:
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)
    except Exception:
        return text

# ── Data ───────────────────────────────────────────────────────────────────────
DAYS = [
    {
        "day": "17.11.2026",
        "title": "נגויה → קיוטו",
        "status": "שינוי",
        "color": "FFF2CC",
        "rows": [
            ("08:20-11:30", "Toyota Kaikan Museum", "ללא שינוי"),
            ("12:00-13:00", "שינקנסן לקיוטו", "ללא שינוי"),
            ("13:00",       "הגעה לתחנת קיוטו — ישר לחקור (ללא מזוודות)", "חדש: Yamato שלח מזוודות מראש"),
            ("13:10-14:20", "Toji Temple — פגודה גבוהה ביפן, UNESCO, ¥500", "חדש"),
            ("14:30-16:00", "Nishiki Market — 'מטבח קיוטו', 400 מ' קיירה", "הועבר מיום 21 (היה ב-08:00)"),
            ("16:00-16:50", "גשר Sanjo-Ohashi / נהר קאמו — מנוחה, קפה", "חדש"),
            ("17:00",       "צ'ק-אין Hotel Monterey Kyoto", "הוזז לשעת צ'ק-אין אמיתית"),
            ("17:30-19:00", "Teramachi Shopping Arcade", "ללא שינוי"),
            ("19:00-21:00", "Pontocho Alley — ארוחת ערב", "ללא שינוי"),
        ],
        "note": "יום 21 נהיה נקי יותר — Nishiki Market מוסר מהבוקר הצפוף של יציאה לאוסקה",
    },
    {
        "day": "18.11.2026",
        "title": "קיוטו — טיול מודרך (6 אתרים)",
        "status": "שינוי סדר בלבד",
        "color": "FCE4D6",
        "rows": [
            ("09:00-10:30", "Arashiyama Bamboo Grove + Tenryu-ji", "קלאסטר מערבי — ביחד, ברגל"),
            ("10:30-11:00", "נסיעה: Arashiyama → Kinkaku-ji (~30 דק' אוטובוס)", ""),
            ("11:00-12:00", "Kinkaku-ji (הפביליון הזהוב)", ""),
            ("12:00-13:00", "Ryoan-ji — גן סלעים זן (10 דק' הליכה מ-Kinkaku-ji)", ""),
            ("13:00-14:00", "ארוחת צהריים", ""),
            ("14:00-14:45", "נסיעה: Ryoan-ji → Fushimi Inari (~45 דק' מטרו+JR)", "נסיעה אחת גדולה"),
            ("14:45-16:30", "Fushimi Inari — אלפי שערי טורי, אור זהב 15:30+", "שינוי: אור אחה\"צ עדיף על בוקר"),
            ("17:00-19:00", "Gion — רובע גייש'ות, שעת ערב", "ללא שינוי"),
        ],
        "note": "אותן 6 אטרקציות כמקור — רק הסדר שונה. מסלול: מערב→צפון-מערב→דרום→מזרח (לעומת דרום→מערב→צפון-מערב→מזרח במקור)",
    },
    {
        "day": "19.11.2026",
        "title": "קיוטו — מקדשים, שביל הפילוסוף, טירה",
        "status": "שינוי + תוספת",
        "color": "E2EFDA",
        "rows": [
            ("09:00-09:45", "Eikando Temple — עלי כסף/זהב, עונת koyo שיא", "ללא שינוי"),
            ("09:50-10:10", "Philosopher's Path (哲学の道) — 2 ק\"מ לאורך תעלה", "חדש — עלים נפולים על המים"),
            ("10:10-11:00", "Nanzen-ji — אקוודוקט + שער ענק, חינם", "חדש"),
            ("11:10-11:50", "Kiyomizu-dera", "ללא שינוי"),
            ("11:50-12:50", "Sannenzaka + Ninenzaka — שעה שלמה", "ללא שינוי (תוכנית מקורית)"),
            ("12:50-13:05", "Yasaka Pagoda", "ללא שינוי"),
            ("13:15-14:15", "ארוחת צהריים", "ללא שינוי"),
            ("14:30-16:00", "Nijo Castle (二条城) — טירת השוגון, UNESCO, 'רצפות לילית'", "חדש — מחליף 4 שעות Shijo Kawaramachi"),
            ("16:30",       "חזרה למלון", ""),
            ("19:00+",      "L'Escamoteur Bar + ארוחת ערב + Turquoise Bar", "ללא שינוי"),
        ],
        "note": "קניות גדולות שמורות לטוקיו (גינזה/שיבויה/שינג'וקו) — Shijo Kawaramachi = רשתות שאפשר למצוא בכל יפן",
    },
    {
        "day": "20.11.2026",
        "title": "נארה — טיול יום",
        "status": "תיקון קטן",
        "color": "DDEBF7",
        "rows": [
            ("07:55",       "יציאה מ-Hotel Monterey Kyoto", "חדש — חסר מהתוכנית המקורית"),
            ("08:20",       "JR Miyakoji Express מקיוטו לנארה (~45 דק')", "חדש — פרט תחבורה"),
            ("09:00-10:00", "Todai-ji + Nigatsu-do Hall (5 דק' מאחורי המקדש)", "תוספת: Nigatsu-do — נוף פנורמי חינם"),
            ("10:00-11:30", "Nara Park — איילים ידידותיים", "ללא שינוי"),
            ("11:30-13:00", "Kasuga Taisha — אלפי פנסי אבן", "ללא שינוי"),
            ("13:00-14:00", "ארוחת צהריים", "ללא שינוי"),
            ("14:00-15:00", "Kofuku-ji — פגודה + מוזיאון", "ללא שינוי"),
            ("15:00-16:00", "Nara Machi — רובע מסחר עתיק", "ללא שינוי"),
            ("16:00-17:00", "חזרה לקיוטו — JR Nara Line", "ללא שינוי"),
            ("19:00-20:30", "ארוחת ערב ב-Pontocho/Kiyamachi", "ללא שינוי"),
        ],
        "note": "יום 20 הוא הכי קרוב לשלם — מסלול גיאוגרפי נכון, בחירת אטרקציות מצוינת",
    },
]

STATUS_COLORS = {
    "שינוי":              "FF0000",
    "שינוי סדר בלבד":    "FF6600",
    "שינוי + תוספת":     "7030A0",
    "תיקון קטן":          "0070C0",
    "ללא שינוי":          "70AD47",
}

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def make_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "הצעות לשיפור קיוטו"
    ws.sheet_view.rightToLeft = True

    # Column widths
    ws.column_dimensions['A'].width = 16   # time
    ws.column_dimensions['B'].width = 42   # activity
    ws.column_dimensions['C'].width = 38   # change note
    ws.column_dimensions['D'].width = 10   # status badge

    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(r, c, val, bold=False, bg=None, color="000000", wrap=True, size=10, align='right'):
        cell_obj = ws.cell(row=r, column=c, value=val)
        cell_obj.font = Font(name='Arial', bold=bold, color=color, size=size)
        cell_obj.alignment = Alignment(horizontal=align, vertical='center',
                                       wrap_text=wrap, readingOrder=2)
        if bg:
            cell_obj.fill = PatternFill("solid", fgColor=bg)
        cell_obj.border = border
        return cell_obj

    row = 1
    # Title row
    ws.merge_cells(f'A{row}:D{row}')
    cell(row, 1, "הצעות לשיפור קיוטו — יפן נובמבר 2026",
         bold=True, bg="1F3864", color="FFFFFF", size=14, align='center')
    ws.row_dimensions[row].height = 28
    row += 1

    for day_data in DAYS:
        # Day header
        ws.merge_cells(f'A{row}:C{row}')
        cell(row, 1, f"{day_data['day']}  |  {day_data['title']}",
             bold=True, bg=day_data['color'], size=11)
        status_color = STATUS_COLORS.get(day_data['status'], "888888")
        cell(row, 4, day_data['status'], bold=True,
             bg=status_color, color="FFFFFF", align='center')
        ws.row_dimensions[row].height = 22
        row += 1

        # Column headers
        for c, txt in enumerate(["שעה", "פעילות / אטרקציה", "הערת שינוי", ""], 1):
            cell(row, c, txt, bold=True, bg="D9D9D9", size=9)
        ws.row_dimensions[row].height = 18
        row += 1

        for time_str, activity, change in day_data['rows']:
            change_bg = "FFF2CC" if change and "חדש" in change else \
                        "DDEBF7" if change and "הועבר" in change else None
            cell(row, 1, time_str, bg="F5F5F5")
            cell(row, 2, activity)
            cell(row, 3, change, bg=change_bg)
            cell(row, 4, "")
            ws.row_dimensions[row].height = 20
            row += 1

        # Note row
        ws.merge_cells(f'A{row}:D{row}')
        cell(row, 1, f"💡 {day_data['note']}", bg="FFF8DC", size=9, color="555555")
        ws.row_dimensions[row].height = 32
        row += 2

    # Legend
    ws.merge_cells(f'A{row}:D{row}')
    cell(row, 1, "מקרא סטטוס שינויים:", bold=True, bg="E0E0E0")
    row += 1
    for status, color in STATUS_COLORS.items():
        ws.merge_cells(f'A{row}:D{row}')
        cell(row, 1, f"  {status}", bg=color, color="FFFFFF", bold=True)
        row += 1

    fname = "הצעות לשיפור קיוטו.xlsx"
    wb.save(fname)
    print(f"Excel saved: {fname}")

# ══════════════════════════════════════════════════════════════════════════════
# PDF  (using Windows Arial which has full Hebrew support)
# ══════════════════════════════════════════════════════════════════════════════
def make_pdf():
    # Register Windows Arial with Hebrew support
    arial_path = r"C:\Windows\Fonts\arial.ttf"
    arialbd_path = r"C:\Windows\Fonts\arialbd.ttf"
    pdfmetrics.registerFont(TTFont('Arial', arial_path))
    pdfmetrics.registerFont(TTFont('Arial-Bold', arialbd_path))

    fname = "הצעות לשיפור קיוטו.pdf"
    doc = SimpleDocTemplate(fname, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    rtl_style = ParagraphStyle('rtl', fontName='Arial', fontSize=10,
                                alignment=TA_RIGHT, leading=14)
    rtl_bold = ParagraphStyle('rtl_bold', fontName='Arial-Bold', fontSize=10,
                               alignment=TA_RIGHT, leading=14)
    title_style = ParagraphStyle('title', fontName='Arial-Bold', fontSize=16,
                                  alignment=TA_CENTER, leading=20,
                                  textColor=colors.white)
    day_style = ParagraphStyle('day', fontName='Arial-Bold', fontSize=11,
                                alignment=TA_RIGHT, leading=16)
    note_style = ParagraphStyle('note', fontName='Arial', fontSize=8,
                                 alignment=TA_RIGHT, textColor=colors.HexColor('#555555'))

    STATUS_PDF_COLORS = {
        "שינוי":              colors.HexColor("#FF0000"),
        "שינוי סדר בלבד":    colors.HexColor("#FF6600"),
        "שינוי + תוספת":     colors.HexColor("#7030A0"),
        "תיקון קטן":          colors.HexColor("#0070C0"),
    }

    CHANGE_TYPE_COLORS = {
        "חדש":   colors.HexColor("#FFF2CC"),
        "הועבר": colors.HexColor("#DDEBF7"),
    }

    story = []

    # Title block
    title_tbl = Table([[Paragraph(h("הצעות לשיפור קיוטו — יפן נובמבר 2026"), title_style)]],
                      colWidths=[17.5*cm])
    title_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#1F3864")),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.HexColor("#1F3864")]),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('ROUNDEDCORNERS', [4]),
    ]))
    story.append(title_tbl)
    story.append(Spacer(1, 0.4*cm))

    col_w = [3*cm, 9*cm, 5.5*cm]

    for day_data in DAYS:
        day_color = colors.HexColor("#" + day_data['color'])
        status_color = STATUS_PDF_COLORS.get(day_data['status'], colors.grey)

        # Day header row
        day_header = [
            Paragraph(h(day_data['status']), ParagraphStyle('sh', fontName='Arial-Bold',
                       fontSize=9, textColor=colors.white, alignment=TA_CENTER)),
            Paragraph(h(f"{day_data['day']}  |  {day_data['title']}"), day_style),
            "",
        ]
        day_tbl = Table([day_header], colWidths=col_w)
        day_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), status_color),
            ('BACKGROUND', (1,0), (2,0), day_color),
            ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ]))
        story.append(day_tbl)

        # Column sub-headers
        hdr_row = [
            Paragraph(h("שעה"), ParagraphStyle('hdr', fontName='Arial-Bold', fontSize=8,
                        alignment=TA_RIGHT)),
            Paragraph(h("פעילות / אטרקציה"), ParagraphStyle('hdr', fontName='Arial-Bold',
                        fontSize=8, alignment=TA_RIGHT)),
            Paragraph(h("הערת שינוי"), ParagraphStyle('hdr', fontName='Arial-Bold',
                        fontSize=8, alignment=TA_RIGHT)),
        ]
        hdr_tbl = Table([hdr_row], colWidths=col_w)
        hdr_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#D9D9D9")),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('LINEAFTER', (0,0), (1,0), 0.5, colors.grey),
        ]))
        story.append(hdr_tbl)

        # Activity rows
        for time_str, activity, change in day_data['rows']:
            # Pick row bg
            row_bg = colors.white
            for key, bg in CHANGE_TYPE_COLORS.items():
                if key in change:
                    row_bg = bg
                    break

            data_row = [
                Paragraph(h(time_str), rtl_style),
                Paragraph(h(activity), rtl_style),
                Paragraph(h(change), ParagraphStyle('chg', fontName='Arial', fontSize=8,
                            alignment=TA_RIGHT, textColor=colors.HexColor("#333333"))),
            ]
            row_tbl = Table([data_row], colWidths=col_w)
            row_tbl.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), row_bg),
                ('TOPPADDING', (0,0), (-1,-1), 3),
                ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                ('LEFTPADDING', (0,0), (-1,-1), 4),
                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
                ('LINEAFTER', (0,0), (1,0), 0.5, colors.HexColor("#CCCCCC")),
            ]))
            story.append(row_tbl)

        # Note
        note_tbl = Table(
            [[Paragraph(h(f"💡  {day_data['note']}"), note_style)]],
            colWidths=[sum(col_w)])
        note_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#FFFACD")),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#CCCC88")),
        ]))
        story.append(note_tbl)
        story.append(Spacer(1, 0.35*cm))

    doc.build(story)
    print(f"PDF saved: {fname}")


if __name__ == "__main__":
    make_excel()
    make_pdf()
    print("Done.")
